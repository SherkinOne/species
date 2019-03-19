import requests
import json
import openpyxl
import sys

def checkSpecies(species, phylum):
    # WORMS is accessed for marine based
    # Global diversity Information network for land
    URL=""
    if species is not None :
        temp=species.split()
        speciesName=""
        for i in temp :
            speciesName+=i+"%20"
    if phylum =="marine" :   
        URL = "http://marinespecies.org/rest/AphiaRecordsByMatchNames?scientificnames%5B%5D="+speciesName+"&marine_only=true"
    if phylum =="land" :
        URL="http://api.gbif.org/v1/species/search?datasetKey=7ddf754f-d193-4cc9-b351-99906754a03b&q="+speciesName
    # call the URL and return relative data
    try:
            response = requests.get(URL)
            if response.status_code==200 :
                todos = json.loads(response.text)
                if phylum=="marine" :
                    if todos[0][0]['scientificname'] is not None :
                        #there is an accepted option to say whether or not the name is accepted
                        return (todos[0][0]['scientificname'],todos[0][0]['valid_name'],todos[0][0]['phylum'])
                else :
                    if todos['results']!=[] :
                        if (todos['results'][0]['taxonomicStatus']).lower()!="accepted" :
                            newname=todos['results'][0]['accepted']
                        else :
                            newname=""
                        return (todos['results'][0]['scientificName'],newname,todos['results'][0]['phylum'])
                    return("","No result","")
            else :
                return("","No result","")
    except :
            print("Quitting")
            wb.save(fileName)
            quit()
    else:
        return ("Finished","","Error")

def searchDatabase(fileName, sheet, phylum):
    try :
        cellValue="*"
        wb = openpyxl.load_workbook(filename=fileName , read_only=False)
        if sheet !="" :
            ws= wb.get_sheet_by_name(sheet)
        else :
            sheets=wb.get_sheet_names()
            ws= wb.get_sheet_by_name(sheets[0])
        counter =0
        scName=""       
        while cellValue is not None or scName!="Finished" :
            counter+=1
            cellValue=ws['A'+str(counter)].value
            scName,valid,phyla=checkSpecies(cellValue, phylum.lower())
            # if scname not same as valid then change - add phylum
            if cellValue.lower()!=valid.lower():
                cell='B'+str(counter)
                ws[cell]=valid
            if phyla!="" :
                ws['C'+str(counter)]=phyla
        print("Saving")
        wb.save(fileName)
    except : 
        print ("Error - saving")
        wb.save(fileName)